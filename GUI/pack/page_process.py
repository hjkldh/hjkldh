import pandas as pd
import os
import xlwt
from xlrd import open_workbook
from xlutils.copy import copy
import time
from openpyxl.styles import Alignment, Border, Side
from openpyxl import Workbook,load_workbook
from datetime import datetime, timedelta
from . import Message


def process_class(*args):
    # 读取 Excel 文件
    df = pd.read_excel(args[0], sheet_name='导出的数据')

    # 按照主要依据“日期”降序排序，次要依据“课程”、“姓名”、“节次”升序排序
    df = df.sort_values(by=['日期', '课程', '姓名', '节次'], ascending=[False, True, True, True])

    # 初始化一个字典来存储每个学生的考勤情况
    attendance_dict = {}

    # 遍历每一行数据
    for index, row in df.iterrows():
        student_id = row['学号']
        student_name = row['姓名']
        course = row['课程']
        date = row['日期']
        period = row['节次']
        status = row['考勤状态']

        # 如果学生不在字典中，初始化其考勤记录
        if student_id not in attendance_dict:
            attendance_dict[student_id] = {
                '姓名': student_name,
                '迟到': 0,
                '旷课': 0,
                '请假': 0,
                '早退': 0,
                '迟到记录': {}  # 用于记录迟到的课程和日期
            }

        # 根据考勤状态更新考勤记录
        if status == '迟到':
            # 判断是否在同一天的连续节次中已经记录过迟到
            key = (course, date)  # 只根据课程和日期判断
            if key in attendance_dict[student_id]['迟到记录']:
                # 如果已经记录过迟到，且当前节次与上一个节次是连续的，则不重复计数
                prev_period = attendance_dict[student_id]['迟到记录'][key]
                current_period_num = int(period.split('第')[1].replace('节', ''))
                prev_period_num = int(prev_period.split('第')[1].replace('节', ''))
                if abs(current_period_num - prev_period_num) == 1:
                    continue
            # 更新迟到记录
            attendance_dict[student_id]['迟到记录'][key] = period
            attendance_dict[student_id]['迟到'] += 1
        elif status == '旷课':
            attendance_dict[student_id]['旷课'] += 1
        elif status == '请假':
            attendance_dict[student_id]['请假'] += 1
        elif status == '早退':
            attendance_dict[student_id]['早退'] += 1

    # 将考勤数据转换为 DataFrame
    attendance_list = []
    for student_id, data in attendance_dict.items():
        attendance_list.append({
            '学号': student_id,
            '姓名': data['姓名'],
            '迟到': data['迟到'],
            '旷课': data['旷课'],
            '请假': data['请假'],
            '早退': data['早退']
        })

    attendance_df = pd.DataFrame(attendance_list)

    # 按学号升序排序
    attendance_df = attendance_df.sort_values(by='学号', ascending=True)

    # 保存结果到新的 Excel 文件
    timestamp = time.strftime('%Y-%m-%d %H:%M:%S')
    if len(args[1]) >= 1:
        output_file_path = args[1]+os.sep+'班级考勤表'+f'_{timestamp}.xlsx'
    else:
        output_file_path = os.path.dirname(args[0])+os.sep+'班级考勤表'+f'_{timestamp}.xlsx'
    with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
        attendance_df.to_excel(writer, index=False, sheet_name='考勤统计')

        # 获取 workbook 和 worksheet 对象
        workbook = writer.book
        worksheet = writer.sheets['考勤统计']

        # 设置单元格居中
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # 设置边框
        border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.border = border

    print(f"考勤统计已保存到 {output_file_path}")
    if len(args[1]) >= 1:
        return args[1]+os.sep
    else:
        return os.path.dirname(args[0]+os.sep)

def generate_attendance_sheet(*args):
    # 解析输入日期（兼容带斜杠的格式）
    try:
        input_date_str = args[0].replace("/", "")
        input_date = datetime.strptime(input_date_str, "%Y%m%d")
    except ValueError:
        print("日期格式错误！请确保输入格式为 YYYY/MM/DD 或 YYYYMMDD。")
        return

    # 读取课程考勤数据
    attendance_df = pd.read_excel(args[1], sheet_name='导出的数据')

    # 读取班级名册（动态识别数据起始行）
    try:
        roster_df = pd.read_excel(args[2], sheet_name='sheet1', header=None)
    except:
        try:
            roster_df = pd.read_excel(args[2], sheet_name='导出的数据', header=None)
        except:
            Message.showerror("错误提示", "请将班级名册中的工作表名称改为“sheet1”或者“导出的数据”后重试")
    # 查找包含“学号”和“姓名”的表头行
    header_row = None
    for idx, row in roster_df.iterrows():
        if "学号" in row.values and "姓名" in row.values:
            header_row = idx
            break

    if header_row is None:
        print("班级名册格式错误：未找到包含'学号'和'姓名'的表头行！")
        return

    # 重新读取名册，指定表头行
    try:
        roster_df = pd.read_excel(args[2],sheet_name='sheet1',header=header_row,usecols=["学号", "姓名"]).dropna(subset=["学号", "姓名"])
    except:
        try:
            roster_df = pd.read_excel(args[2], sheet_name='导出的数据', header=header_row, usecols=["学号", "姓名"]).dropna(subset=["学号", "姓名"])
        except:
            Message.showerror("错误提示", "请将班级名册中的工作表名称改为“sheet1”或者“导出的数据”后重试")
    students = roster_df[["学号", "姓名"]].values.tolist()

    # 创建新的 Excel 文件
    wb = Workbook()
    ws = wb.active
    ws.title = "考勤统计"

    # 设置表头
    headers = ["序号", "学号", "姓名"] + [f"{i}周" for i in range(1, 19)]
    ws.append(headers)

    # 设置列宽
    for col in ws.columns:
        col_letter = col[0].column_letter
        if col[0].value in ["1周", "2周", "3周", "4周", "5周", "6周", "7周", "8周", "9周", "10周", "11周", "12周", "13周", "14周", "15周", "16周", "17周", "18周"]:
            ws.column_dimensions[col_letter].width = 4
        else:
            ws.column_dimensions[col_letter].width = 10

    # 设置单元格居中和边框
    alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    # 填充学生名单
    for idx, (student_id, student_name) in enumerate(students, start=1):
        row = [idx, student_id, student_name] + [""] * 18
        ws.append(row)

    # 计算教学周的起始日期（假设输入的日期是某周的周一）
    start_date = input_date - timedelta(days=input_date.weekday())

    # 初始化考勤字典：{学号: {周次: 符号}}
    attendance_dict = {student_id: {week: "" for week in range(1, 19)} for student_id, _ in students}

    # 处理考勤数据
    attendance_df = attendance_df.sort_values(by=["日期", "课程", "姓名", "节次"], ascending=[True, True, True, True])

    for _, row in attendance_df.iterrows():
        student_id = row["学号"]
        date = pd.to_datetime(row["日期"]).date()
        status = row["考勤状态"]
        period = row["节次"]

        # 计算周次
        delta_days = (date - start_date.date()).days
        if delta_days < 0:
            continue  # 忽略早于起始日期的记录
        week_number = (delta_days // 7) + 1
        if week_number > 18:
            continue

        # 处理迟到合并逻辑
        if status == "迟到":
            course = row["课程"]
            date_key = (student_id, course, date)
            current_period_num = int(period.split("第")[1].replace("节", ""))

            # 检查是否已有同课程、同一天的连续节次迟到
            prev_period = attendance_dict.get("_late_records", {}).get(date_key)
            if prev_period:
                prev_period_num = int(prev_period.split("第")[1].replace("节", ""))
                if abs(current_period_num - prev_period_num) == 1:
                    continue  # 连续节次迟到，跳过
            attendance_dict.setdefault("_late_records", {})[date_key] = period

        # 映射符号
        symbol_map = {
            "公假": "◎", "病假": "○", "事假": "△",
            "迟到": "#", "早退": "*", "旷课": "×"
        }
        symbol = symbol_map.get(status, "")

        # 更新考勤记录（覆盖写入，确保最新状态）
        if student_id in attendance_dict:
            attendance_dict[student_id][week_number] = symbol

    # 填充考勤到 Excel
    for row_idx in range(2, len(students) + 2):  # 从第2行开始
        student_id = ws.cell(row=row_idx, column=2).value
        for week in range(1, 19):
            col_idx = 3 + week  # 第4列对应"1周"
            symbol = attendance_dict.get(student_id, {}).get(week, "")
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = symbol

    # 遍历所有行（从第一行开始，包括表头）
    for row in ws.iter_rows(min_row=1, max_row=len(students) + 1, min_col=1, max_col=21):
        for cell in row:
            cell.alignment = alignment
            cell.border = border
    # 保存文件
    timestamp = time.strftime('%Y-%m-%d %H:%M:%S')
    if len(args[3]) >= 1:
        output_file_path = args[3] + os.sep + '课程考勤表' + f'_{timestamp}.xlsx'
    else:
        output_file_path = os.path.dirname(args[1]) + os.sep + '课程考勤表' + f'_{timestamp}.xlsx'
    wb.save(output_file_path)
    print(f"考勤统计表已生成：{output_file_path}")
    if len(args[3]) >= 1:
        return args[3]+os.sep
    else:
        return os.path.dirname(args[1])+os.sep


def calculate_grades(
    roster_path: str,         # 班级名册路径
    attendance_path: str,     # 考勤记录路径
    q_values: list,           # [Q1, Q2, Q3, Q4] 总和需为100
    a_values: list,           # [A1, A2, A3, A4, A5, A6]
    homework_path: str = None,# 作业成绩路径（可选）
    exam_path: str = None,    # 期末成绩路径
    output_dir: str = "."     # 输出目录
) -> str:
    """
    生成学生成绩统计表算法
    """
    # --------------------------
    # 1. 参数校验
    # --------------------------
    if sum(q_values) != 100:
        raise ValueError("成绩占比总和必须为 100")
    if len(a_values) != 6:
        raise ValueError("扣分值需全部输入")
    symbol_map = {
        "◎": a_values[0], "○": a_values[1],
        "△": a_values[2], "#": a_values[3],
        "*": a_values[4], "×": a_values[5]
    }

    # --------------------------
    # 2. 读取班级名册
    # --------------------------
    def find_header(df, keywords=["学号", "姓名"]):
        """动态查找包含关键字的表头行"""
        for idx, row in df.iterrows():
            if all(col in row.values for col in keywords):
                return idx
        return None

    # 首次读取用于定位表头
    roster_df_raw = pd.read_excel(roster_path, header=None)
    header_row = find_header(roster_df_raw)
    if header_row is None:
        Message.showerror("错误提示","班级名册中未找到学号/姓名列")

    # 正式读取数据
    roster_df = pd.read_excel(
        roster_path,
        header=header_row,
        usecols=["学号", "姓名"]
    ).dropna(subset=["学号", "姓名"])
    students = roster_df[["学号", "姓名"]].values.tolist()

    # --------------------------
    # 3. 处理考勤记录 (Q1) - 修复核心逻辑
    # --------------------------
    # 读取考勤表结构（假设列名格式：学号、姓名、1周、2周...18周）
    attendance_df_raw = pd.read_excel(attendance_path, header=None)

    # 动态识别表头
    week_columns = []
    for idx, row in attendance_df_raw.iterrows():
        if "学号" in row.values and "姓名" in row.values:
            header_row = idx
            # 识别周次列（1周 至 18周）
            week_columns = [col for col in attendance_df_raw.iloc[idx] if isinstance(col, str) and "周" in col]
            break

    if not week_columns:
        raise ValueError("考勤表中未找到周次列（如'1周'）")

    # 重新读取考勤表
    attendance_df = pd.read_excel(
        attendance_path,
        header=header_row,
        usecols=["学号", "姓名"] + week_columns
    )

    # 计算每个学生的扣分总和
    Q1_scores = {}
    Q1_max = q_values[0]
    for _, row in attendance_df.iterrows():
        student_id = row["学号"]
        total_deduction = 0

        # 遍历所有周次列
        for week_col in week_columns:
            symbol = row.get(week_col, "")
            if symbol in symbol_map:
                total_deduction += symbol_map[symbol]

        # 计算最终得分
        final_score = max(Q1_max - total_deduction, 0)
        Q1_scores[student_id] = final_score

    # --------------------------
    # 4. 处理作业成绩 (Q3)
    # --------------------------
    Q3_scores = {}
    if homework_path:
        # 读取作业成绩
        hw_df_raw = pd.read_excel(homework_path, header=None)
        hw_header = find_header(hw_df_raw, ["学号", "成绩"])
        if hw_header is None:
            raise ValueError("作业成绩文件格式错误")

        hw_df = pd.read_excel(
            homework_path,
            header=hw_header,
            usecols=["学号", "成绩"]
        )

        # 检查是否需要折算：只要有一个学生分数超过 Q3，所有人按比例折算
        max_hw = hw_df["成绩"].max()
        Q3_max = q_values[2]
        scale = Q3_max / 100 if max_hw > Q3_max else 1.0
        for _, row in hw_df.iterrows():
            student_id = row["学号"]
            raw_score = row["成绩"]
            Q3_scores[student_id] = round(raw_score * scale, 1)
    else:
        # 统一赋值
        default_score = q_values[2]
        Q3_scores = {student_id: default_score for student_id, _ in students}

    # --------------------------
    # 5. 处理期末成绩 (Q4)
    # --------------------------
    Q4_scores = {}
    # 读取期末成绩
    exam_df_raw = pd.read_excel(exam_path, header=None)
    exam_header = find_header(exam_df_raw, ["学号", "成绩"])
    if exam_header is None:
        raise ValueError("期末成绩文件格式错误")

    exam_df = pd.read_excel(exam_path,header=exam_header,usecols=["学号", "成绩"])

    # 检查是否需要折算：只要有一个学生分数超过 Q4，所有人按比例折算
    max_exam = exam_df["成绩"].max()
    Q4_max = q_values[3]
    if max_exam > Q4_max:
        scale = Q4_max / 100
        for _, row in exam_df.iterrows():
            student_id = row["学号"]
            raw_score = row["成绩"]
            Q4_scores[student_id] = round(raw_score * scale, 1)
    else:
        # 统一赋值
        default_score = q_values[3]
        Q4_scores = {student_id: default_score for student_id, _ in students}

    # --------------------------
    # 6. 整合数据并生成 Excel
    # --------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "成绩统计"

    # 设置表头
    headers = [
        "序号", "学号", "姓名",
        f"缺勤记录（{q_values[0]}分）",
        f"课堂参与（{q_values[1]}分）",
        f"作业（{q_values[2]}分）",
        f"期末测试（{q_values[3]}分）",
        "合计（100分）"
    ]
    ws.append(headers)

    # 设置样式
    alignment = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # 填充数据
    for idx, (student_id, name) in enumerate(students, start=1):
        q1 = Q1_scores.get(student_id, 0)
        q2 = q_values[1]  # 课堂参与直接赋值
        q3 = Q3_scores.get(student_id, 0)
        q4 = Q4_scores.get(student_id, 0)
        total = q1 + q2 + q3 + q4

        row = [idx, student_id, name, q1, q2, q3, q4, total]
        ws.append(row)

    # 应用样式至所有单元格
    for row in ws.iter_rows(min_row=1, max_row=len(students)+1):
        for cell in row:
            cell.alignment = alignment
            cell.border = border

    # 自动调整列宽
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[col[0].column_letter].width = adjusted_width

    # 保存文件
    timestamp = time.strftime('%Y-%m-%d %H:%M:%S')
    if output_dir:
        output_path = os.path.join(output_dir, f"考核记录单_{timestamp}.xlsx")
    else:
        output_path = os.path.join(os.path.dirname(attendance_path),f"考核记录单_{timestamp}.xlsx")
    wb.save(output_path)

    return output_path


def find_column(df, possible_names):
    """智能查找列名"""
    cols = df.columns.str.strip().str.lower()  # 标准化列名
    for name in possible_names:
        clean_name = name.strip().lower()
        if clean_name in cols:
            return df.columns[cols == clean_name][0]
    raise KeyError(f"未找到匹配列，尝试的列名：{possible_names}")

def fill_scores(template_path, score_path, midterm_path, weights, save_dir=None):
    # 参数校验
    if sum(weights) != 100:
        Message.showerror("错误提示","成绩占比总和必须为 100")
    #weights = [n/100 for n in weights]
    B1, B2, B3, B4, B5 = weights
    if B2 != 0 and not midterm_path:
        Message.showerror("错误提示","期中成绩比重不为0，请传入期中成绩表")

    # 读取数据
    score_df = pd.read_excel(score_path)
    try:
        template_df = pd.read_excel(template_path,sheet_name='导出的数据',engine='xlrd')
    except ImportError:
        raise RuntimeError("读取.xls需要安装xlrd，请执行：pip install xlrd")

    # 动态匹配成绩统计表列名
    col_mapping = {
        'final_exam': ['期末测试（40分）','期末测试（50分）','期末测试（60分）', '期末测试', '期末成绩', 'final exam'],
        'attendance': ['缺勤记录（10分）','缺勤记录（20分）','缺勤记录（30分）', '缺勤记录', '考勤扣分', 'attendance'],
        'participation': ['课堂参与（10分）','课堂参与（20分）','课堂参与（30分）', '课堂参与', '课堂表现', 'participation'],
        'homework': ['作业（0分）','作业（10分）','作业（20分）','作业（30分）','作业（40分）', '作业成绩', '平时作业', 'homework']
    }

    try:
        # 获取实际列名
        final_col = find_column(score_df, col_mapping['final_exam'])
        attend_col = find_column(score_df, col_mapping['attendance'])
        partic_col = find_column(score_df, col_mapping['participation'])
        hw_col = find_column(score_df, col_mapping['homework'])
    except:
        Message.showerror("错误提示","请确保比例设置和“考勤记录表”中一致")

    # 合并数据（保持模板顺序）
    merged_df = pd.merge(
        template_df[['学号']],
        score_df[[final_col, attend_col, partic_col, hw_col, '学号']],
        on='学号',
        how='left'
    )

    # 处理各列成绩
    template_df['期末成绩'] = merged_df[final_col] / B1 * 100
    template_df['考勤'] = merged_df[attend_col] / B4 * 100
    template_df['平时成绩'] = merged_df[partic_col] / B3 * 100
    template_df['作业'] = 0 if B5 == 0 else merged_df[hw_col] / B5 * 100

    # 处理期中成绩（需要单独匹配列名）
    if B2 != 0 and midterm_path:
        midterm_df = pd.read_excel(midterm_path)
        midterm_col = find_column(midterm_df, ['期中成绩', '期中测试', 'midterm'])
        midterm_merged = pd.merge(
            template_df[['学号']],
            midterm_df[['学号', midterm_col]],
            on='学号',
            how='left'
        )
        max_score = midterm_merged[midterm_col].max()
        template_df['期中成绩'] = (
            midterm_merged[midterm_col] if max_score > B2
            else midterm_merged[midterm_col] / B2 * 100
        )
    else:
        template_df['期中成绩'] = 0

    # 计算总评成绩
    template_df['总评成绩'] = (
                                      template_df['期末成绩'] * B1 +
                                      template_df['期中成绩'] * B2 +
                                      template_df['平时成绩'] * B3 +
                                      template_df['考勤'] * B4 +
                                      template_df['作业'] * B5
                              ) / 100

    # 保存结果
    timestamp = time.strftime('%Y-%m-%d %H:%M:%S')
    if save_dir:
        output_path = os.path.join(save_dir, f"上传模板生成成绩_{timestamp}.xls")
    else:
        output_path = os.path.join(os.path.dirname(template_path), f"考核记录单_{timestamp}.xls")

    workbook = xlwt.Workbook()
    worksheet = workbook.add_sheet('导出的数据')
    #设置样式
    align = xlwt.Alignment()
    align.vert = 1
    align.horz = 2
    ft1 = xlwt.Font()
    ft1.height = 12 * 20
    ft1.bold = True
    ft2 = xlwt.Font()
    ft2.height = 10 * 20
    boder = xlwt.Borders()
    boder.top = 1
    boder.bottom = 1
    boder.left = 1
    boder.right = 1
    style1 = xlwt.XFStyle()
    style1.alignment = align
    style1.font = ft1
    style1.borders = boder
    style2 = xlwt.XFStyle()
    style2.alignment = align
    style1.font = ft1
    style2.borders = boder
    worksheet.col(0).width=22 * 256
    worksheet.col(1).width=17 * 256
    # 写入标题行
    headers = template_df.columns.tolist()
    for col_num, header in enumerate(headers):
        worksheet.write(0, col_num, header,style1)

    # 写入数据行（使用生成器提高内存效率）
    for row_num, row_data in enumerate(template_df.itertuples(index=False), 1):
        for col_num, cell_data in enumerate(row_data):
            # 处理特殊数据类型
            if isinstance(cell_data, datetime):
                cell_data = cell_data.strftime('%Y-%m-%d')
                worksheet.write(row_num, col_num, cell_data,style2)
            else:
                worksheet.write(row_num, col_num, cell_data,style2)
    workbook.save(output_path)
    return os.path.dirname(output_path)
